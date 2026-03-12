"""
Скрипт сопоставления хостов к подсетям по правилам сетевого взаимодействия.
Использует принадлежность IP-адреса к CIDR-подсети.

Требования:
    pip install pandas openpyxl

Использование:
    python match_host_subnet.py

Ожидаемые входные файлы:
    - subnet.xlsx  — колонка 'subnet' с CIDR (например: 10.10.0.0/24)
    - host.xlsx    — колонка 'host' с IP-адресами (например: 10.10.0.5)

Результат:
    - result.xlsx  — три колонки: host | subnet | match_status
"""

import ipaddress
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Настройки файлов ──────────────────────────────────────────────────────────
SUBNET_FILE = "subnet.xlsx"   # файл с подсетями
HOST_FILE   = "host.xlsx"     # файл с хостами
OUTPUT_FILE = "result.xlsx"   # результирующий файл

SUBNET_COL  = "subnet"        # название колонки в subnet.xlsx
HOST_COL    = "host"          # название колонки в host.xlsx
# ─────────────────────────────────────────────────────────────────────────────


def load_data(filepath: str, column: str) -> list[str]:
    """Читает xlsx и возвращает список значений из указанной колонки."""
    df = pd.read_excel(filepath, dtype=str)
    df.columns = df.columns.str.strip()
    if column not in df.columns:
        raise ValueError(f"Колонка '{column}' не найдена в {filepath}. "
                         f"Доступные: {list(df.columns)}")
    return df[column].dropna().str.strip().tolist()


def find_subnet(ip_str: str, networks: list) -> str | None:
    """Возвращает наиболее узкую подсеть, которой принадлежит IP."""
    try:
        ip = ipaddress.ip_address(ip_str)
    except ValueError:
        return None

    matched = [net for net in networks if ip in net]
    if not matched:
        return None
    # Выбираем самую специфичную (наибольший prefixlen)
    return str(max(matched, key=lambda n: n.prefixlen))


def parse_networks(subnet_list: list[str]) -> list:
    networks = []
    for s in subnet_list:
        try:
            networks.append(ipaddress.ip_network(s, strict=False))
        except ValueError:
            print(f"  [WARN] Не удалось разобрать подсеть: '{s}' — пропускаем")
    return networks


def build_result(hosts: list[str], networks: list, subnet_strs: list[str]) -> pd.DataFrame:
    rows = []
    for host in hosts:
        subnet = find_subnet(host, networks)
        status = "✓ найдено" if subnet else "✗ не найдено"
        rows.append({"host": host, "subnet": subnet or "", "status": status})
    return pd.DataFrame(rows)


def style_sheet(ws):
    """Применяет профессиональное форматирование к листу."""
    # Цвета
    GREEN  = "4CAF50"
    RED    = "F44336"
    HEADER = "2E7D32"
    WHITE  = "FFFFFF"
    LIGHT  = "F5F5F5"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовки
    headers = {"A1": "Host", "B1": "Subnet", "C1": "Статус"}
    for cell_addr, title in headers.items():
        cell = ws[cell_addr]
        cell.value = title
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", start_color=HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Данные
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill_color = LIGHT if row_idx % 2 == 0 else WHITE
        status_val = ws.cell(row=row_idx, column=3).value or ""

        for col_idx, cell in enumerate(row, start=1):
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

            if col_idx == 3:
                # Цвет статуса
                if "✓" in status_val:
                    cell.fill = PatternFill("solid", start_color="C8E6C9")  # светло-зелёный
                    cell.font = Font(name="Arial", size=10, color=GREEN)
                else:
                    cell.fill = PatternFill("solid", start_color="FFCDD2")  # светло-красный
                    cell.font = Font(name="Arial", size=10, color=RED)
            else:
                cell.fill = PatternFill("solid", start_color=fill_color)

    # Ширина колонок
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.row_dimensions[1].height = 24


def main():
    print("=== Сопоставление хостов к подсетям ===\n")

    print(f"[1/4] Загрузка подсетей из '{SUBNET_FILE}'...")
    subnet_strs = load_data(SUBNET_FILE, SUBNET_COL)
    networks = parse_networks(subnet_strs)
    print(f"      Загружено подсетей: {len(networks)}")

    print(f"[2/4] Загрузка хостов из '{HOST_FILE}'...")
    hosts = load_data(HOST_FILE, HOST_COL)
    print(f"      Загружено хостов: {len(hosts)}")

    print("[3/4] Сопоставление IP → подсеть...")
    df = build_result(hosts, networks, subnet_strs)
    matched = (df["status"].str.startswith("✓")).sum()
    print(f"      Сопоставлено: {matched}/{len(hosts)}")

    print(f"[4/4] Сохранение результата в '{OUTPUT_FILE}'...")
    df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    ws.title = "Результат"
    style_sheet(ws)

    # Лист статистики
    ws_stat = wb.create_sheet("Статистика")
    ws_stat["A1"] = "Метрика"
    ws_stat["B1"] = "Значение"
    stats = [
        ("Всего хостов", len(hosts)),
        ("Всего подсетей", len(networks)),
        ("Сопоставлено", int(matched)),
        ("Не найдено", int(len(hosts) - matched)),
    ]
    for i, (k, v) in enumerate(stats, start=2):
        ws_stat[f"A{i}"] = k
        ws_stat[f"B{i}"] = v

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Готово! Файл сохранён: {OUTPUT_FILE}")
    print(f"  Лист 'Результат'   — полная таблица сопоставлений")
    print(f"  Лист 'Статистика'  — сводка")


if __name__ == "__main__":
    main()
