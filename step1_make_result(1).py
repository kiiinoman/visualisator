"""
Шаг 1. Сопоставление хостов к подсетям + уникальные хосты в отдельный лист.

Логика:
  1. Читает subnet.xlsx  — колонка 'subnet' (CIDR, например 10.10.0.0/24)
  2. Читает host.xlsx    — колонка 'host'   (IP-адреса, например 10.10.0.5)
  3. Для каждого хоста ищет подходящую подсеть (наиболее специфичную).
  4. Сохраняет result.xlsx с двумя листами:
       - 'result'           — host | subnet | статус  (все хосты)
       - 'Уникальные хосты' — хосты, для которых подсеть НЕ найдена

Требования:
    pip install pandas openpyxl

Входные файлы (рядом со скриптом):
    - subnet.xlsx
    - host.xlsx

Результат:
    - result.xlsx
"""

from typing import Optional, List
import ipaddress
import collections

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Настройки ─────────────────────────────────────────────────────────────────
SUBNET_FILE = "subnet.xlsx"
HOST_FILE   = "host.xlsx"
OUTPUT_FILE = "result.xlsx"

SUBNET_COL  = "subnet"   # заголовок колонки в subnet.xlsx
HOST_COL    = "host"     # заголовок колонки в host.xlsx
# ─────────────────────────────────────────────────────────────────────────────


# ── Стили ─────────────────────────────────────────────────────────────────────
COLOR_HEADER  = "2E7D32"
COLOR_HEADER2 = "1565C0"
COLOR_MATCHED = "C8E6C9"
COLOR_UNIQUE  = "FFCDD2"
COLOR_WHITE   = "FFFFFF"
COLOR_LIGHT   = "F5F5F5"


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(cell, text, bg=COLOR_HEADER):
    cell.value = text
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin_border()
    return cell


def _data_cell(cell, bg=COLOR_WHITE):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _thin_border()
    return cell
# ─────────────────────────────────────────────────────────────────────────────


def load_column(filepath, column):
    # type: (str, str) -> List[str]
    df = pd.read_excel(filepath, dtype=str)
    df.columns = df.columns.str.strip()
    # поиск без учёта регистра
    col_map = {c.lower(): c for c in df.columns}
    matched = col_map.get(column.lower())
    if not matched:
        raise ValueError(
            "Колонка '%s' не найдена в %s. Доступные: %s"
            % (column, filepath, list(df.columns))
        )
    return df[matched].dropna().str.strip().tolist()


def parse_networks(subnet_list):
    # type: (List[str]) -> List
    networks = []
    for s in subnet_list:
        try:
            networks.append(ipaddress.ip_network(s, strict=False))
        except ValueError:
            print("  [WARN] Не удалось разобрать подсеть: '%s' — пропускаем" % s)
    return networks


def find_best_subnet(ip_str, networks):
    # type: (str, list) -> Optional[str]
    """Возвращает наиболее специфичную подсеть для IP или None."""
    try:
        ip = ipaddress.ip_address(ip_str)
    except ValueError:
        return None
    matched = [net for net in networks if ip in net]
    if not matched:
        return None
    return str(max(matched, key=lambda n: n.prefixlen))


def write_result_sheet(ws, rows):
    """
    Записывает лист 'result'.
    rows: list of (host, subnet, status)
    """
    # Заголовки
    for col_idx, title in enumerate(["host", "subnet", "статус"], start=1):
        _header_cell(ws.cell(row=1, column=col_idx), title)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.row_dimensions[1].height = 22

    for row_idx, (host, subnet, status) in enumerate(rows, start=2):
        bg_row = COLOR_MATCHED if subnet else COLOR_UNIQUE
        _data_cell(ws.cell(row=row_idx, column=1, value=host),   bg=bg_row)
        _data_cell(ws.cell(row=row_idx, column=2, value=subnet),  bg=bg_row)
        # Статус с цветом текста
        c = ws.cell(row=row_idx, column=3, value=status)
        _data_cell(c, bg=bg_row)
        c.font = Font(
            name="Arial", size=10,
            color="4CAF50" if subnet else "F44336"
        )


def write_unique_sheet(ws, unique_hosts):
    # type: (object, List[str]) -> None
    """Записывает лист 'Уникальные хосты'."""
    _header_cell(ws.cell(row=1, column=1), "host (подсеть не найдена)", bg=COLOR_HEADER2)
    ws.column_dimensions["A"].width = 25
    ws.row_dimensions[1].height = 22

    for row_idx, host in enumerate(unique_hosts, start=2):
        bg = COLOR_LIGHT if row_idx % 2 == 0 else COLOR_WHITE
        _data_cell(ws.cell(row=row_idx, column=1, value=host), bg=bg)



def write_unique_subnets_sheet(ws, unique_subnets):
    # type: (object, List[str]) -> None
    """Записывает лист 'Уникальные подсети'."""
    _header_cell(ws.cell(row=1, column=1), "subnet (нет совпадений с хостами)", bg=COLOR_HEADER2)
    ws.column_dimensions["A"].width = 28
    ws.row_dimensions[1].height = 22

    for row_idx, subnet in enumerate(unique_subnets, start=2):
        bg = COLOR_LIGHT if row_idx % 2 == 0 else COLOR_WHITE
        _data_cell(ws.cell(row=row_idx, column=1, value=subnet), bg=bg)

def main():
    print("=== Шаг 1: Сопоставление хостов к подсетям ===\n")

    print("[1/4] Загрузка подсетей из '%s'..." % SUBNET_FILE)
    subnet_strs = load_column(SUBNET_FILE, SUBNET_COL)
    networks = parse_networks(subnet_strs)
    print("      Загружено подсетей: %d" % len(networks))

    print("[2/4] Загрузка хостов из '%s'..." % HOST_FILE)
    hosts = load_column(HOST_FILE, HOST_COL)
    print("      Загружено хостов: %d" % len(hosts))

    print("[3/4] Сопоставление IP -> подсеть...")
    rows = []
    unique_hosts = []
    matched_count = 0

    for host in hosts:
        subnet = find_best_subnet(host, networks)
        if subnet:
            status = "найдено"
            matched_count += 1
        else:
            status = "не найдено"
            unique_hosts.append(host)
        rows.append((host, subnet or "", status))

    # Подсети, у которых нет ни одного совпавшего хоста
    matched_subnets = set(subnet for _, subnet, status in rows if subnet)
    unique_subnets = [s for s in subnet_strs if str(ipaddress.ip_network(s, strict=False)) not in matched_subnets]

    print("      Сопоставлено: %d / %d" % (matched_count, len(hosts)))
    print("      Уникальных хостов (без подсети): %d" % len(unique_hosts))
    print("      Уникальных подсетей (без хостов): %d" % len(unique_subnets))

    print("[4/4] Сохранение в '%s'..." % OUTPUT_FILE)

    # Сохраняем через pandas для базовой структуры
    df = pd.DataFrame(rows, columns=["host", "subnet", "статус"])
    df.to_excel(OUTPUT_FILE, index=False, sheet_name="result")

    # Открываем openpyxl для форматирования и добавления листа
    wb = load_workbook(OUTPUT_FILE)
    ws_result = wb["result"]
    write_result_sheet(ws_result, rows)

    ws_unique = wb.create_sheet("Уникальные хосты")
    write_unique_sheet(ws_unique, unique_hosts)

    ws_unique_sn = wb.create_sheet("Уникальные подсети")
    write_unique_subnets_sheet(ws_unique_sn, unique_subnets)

    wb.save(OUTPUT_FILE)

    print("\n✓ Готово! Файл: %s" % OUTPUT_FILE)
    print("  Лист 'result'           — все хосты (%d строк)" % len(rows))
    print("  Лист 'Уникальные хосты'   — без совпадений (%d строк)" % len(unique_hosts))
    print("  Лист 'Уникальные подсети' — без хостов (%d строк)" % len(unique_subnets))


if __name__ == "__main__":
    main()
