"""
Итоговый скрипт сопоставления хостов и подсетей.

Входные файлы:
    subnet.xlsx  — одна колонка 'subnet' (CIDR, например 10.178.0.0/24)
    host.xlsx    — одна колонка 'host'   (IP, например 10.178.25.112)

Результат — result.xlsx с тремя листами:
    'subnet_hosts'   — подсеть | хосты (через запятую) | кол-во хостов
    'unique_hosts'   — хосты, не вошедшие ни в одну подсеть
    'unique_subnets' — подсети, к которым не относится ни один хост

Лог пишется в run.log рядом со скриптом.

Python 3.9+, зависимости: pip install pandas openpyxl
"""

import ipaddress
import logging
import sys
import os
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Файлы ─────────────────────────────────────────────────────────────────────
SUBNET_FILE = "subnet.xlsx"
HOST_FILE   = "host.xlsx"
OUTPUT_FILE = "result.xlsx"
LOG_FILE    = "run.log"
# ─────────────────────────────────────────────────────────────────────────────

# ── Цвета ─────────────────────────────────────────────────────────────────────
C_HDR_GREEN = "2E7D32"
C_HDR_BLUE  = "1565C0"
C_ROW_EVEN  = "F5F5F5"
C_ROW_ODD   = "FFFFFF"
C_HAS_HOSTS = "C8E6C9"
C_NO_HOSTS  = "FFCDD2"
# ─────────────────────────────────────────────────────────────────────────────


# ── Логирование ───────────────────────────────────────────────────────────────
def setup_logger():
    log = logging.getLogger("make_result")
    log.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        fmt="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # В файл — DEBUG и выше (всё)
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8", mode="w")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    # В консоль — INFO и выше
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    log.addHandler(fh)
    log.addHandler(ch)
    return log
# ─────────────────────────────────────────────────────────────────────────────


def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(cell, text, bg=C_HDR_GREEN):
    cell.value = text
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()


def dat(cell, value=None, bg=C_ROW_ODD, bold=False):
    if value is not None:
        cell.value = value
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border()


def read_col(filepath, colname, log):
    # type: (str, str, logging.Logger) -> List[str]
    log.info("Читаем файл: %s (колонка: %s)" % (filepath, colname))
    if not os.path.exists(filepath):
        log.error("Файл не найден: %s" % filepath)
        raise FileNotFoundError("Файл не найден: %s" % filepath)

    df = pd.read_excel(filepath, dtype=str)
    df.columns = df.columns.str.strip()
    log.debug("Колонки в %s: %s" % (filepath, list(df.columns)))

    mapping = {c.lower(): c for c in df.columns}
    real = mapping.get(colname.lower())
    if not real:
        log.error("Колонка '%s' не найдена в %s" % (colname, filepath))
        raise ValueError("Колонка '%s' не найдена в %s. Есть: %s" % (
            colname, filepath, list(df.columns)))

    values = df[real].dropna().str.strip().tolist()
    log.info("Прочитано строк: %d" % len(values))
    log.debug("Первые 5 значений: %s" % values[:5])
    return values


def norm(cidr):
    # type: (str) -> Optional[str]
    try:
        return str(ipaddress.ip_network(cidr, strict=False))
    except ValueError:
        return None


def main():
    log = setup_logger()
    log.info("=" * 60)
    log.info("Старт  %s" % datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("=" * 60)

    # ── 1. Подсети ────────────────────────────────────────────────────────────
    log.info("ШАГ 1/4 — Загрузка подсетей")
    raw_subnets = read_col(SUBNET_FILE, "subnet", log)
    log.info("Строк в файле (raw): %d" % len(raw_subnets))

    seen = set()
    subnets = []
    networks = []
    bad_subnets = []

    for s in raw_subnets:
        n = norm(s)
        if n is None:
            bad_subnets.append(s)
            log.warning("Не удалось разобрать подсеть: '%s' — пропускаем" % s)
        elif n in seen:
            log.debug("Дубликат подсети пропущен: %s" % n)
        else:
            seen.add(n)
            subnets.append(n)
            networks.append(ipaddress.ip_network(n))

    log.info("Уникальных подсетей после дедупликации: %d" % len(subnets))
    log.info("Дублей удалено: %d" % (len(raw_subnets) - len(bad_subnets) - len(subnets)))
    if bad_subnets:
        log.warning("Нераспознанных подсетей: %d -> %s" % (len(bad_subnets), bad_subnets[:10]))

    # ── 2. Хосты ─────────────────────────────────────────────────────────────
    log.info("ШАГ 2/4 — Загрузка хостов")
    raw_hosts = read_col(HOST_FILE, "host", log)
    log.info("Строк в файле (raw): %d" % len(raw_hosts))

    hosts = list(dict.fromkeys(raw_hosts))
    dup_hosts = len(raw_hosts) - len(hosts)
    log.info("Уникальных хостов после дедупликации: %d" % len(hosts))
    if dup_hosts:
        log.info("Дублей хостов удалено: %d" % dup_hosts)

    bad_ip_hosts = []
    for h in hosts:
        try:
            ipaddress.ip_address(h)
        except ValueError:
            bad_ip_hosts.append(h)
    if bad_ip_hosts:
        log.warning("Невалидных IP в host.xlsx: %d -> %s" % (len(bad_ip_hosts), bad_ip_hosts[:10]))

    # ── 3. Сопоставление ─────────────────────────────────────────────────────
    log.info("ШАГ 3/4 — Сопоставление хостов к подсетям")

    subnet_to_hosts = {s: [] for s in subnets}  # type: Dict[str, List[str]]
    unique_hosts = []

    for host in hosts:
        try:
            ip = ipaddress.ip_address(host)
        except ValueError:
            log.debug("Невалидный IP, пропускаем: %s" % host)
            unique_hosts.append(host)
            continue

        best = None
        best_prefix = -1
        for net, cidr_str in zip(networks, subnets):
            if ip in net and net.prefixlen > best_prefix:
                best = cidr_str
                best_prefix = net.prefixlen

        if best:
            subnet_to_hosts[best].append(host)
            log.debug("HOST %s -> SUBNET %s" % (host, best))
        else:
            unique_hosts.append(host)
            log.debug("HOST %s -> не найдена подсеть" % host)

    matched_subnets = [s for s in subnets if subnet_to_hosts[s]]
    unique_subnets  = [s for s in subnets if not subnet_to_hosts[s]]

    log.info("Результат сопоставления:")
    log.info("  Подсетей всего:              %d" % len(subnets))
    log.info("  Подсетей с хостами:          %d" % len(matched_subnets))
    log.info("  Подсетей без хостов:         %d" % len(unique_subnets))
    log.info("  Хостов всего:                %d" % len(hosts))
    log.info("  Хостов сопоставлено:         %d" % (len(hosts) - len(unique_hosts)))
    log.info("  Хостов без подсети:          %d" % len(unique_hosts))

    # Топ подсетей по кол-ву хостов
    top = sorted(matched_subnets, key=lambda s: len(subnet_to_hosts[s]), reverse=True)[:10]
    log.info("Топ-10 подсетей по кол-ву хостов:")
    for s in top:
        log.info("  %-25s -> %d хостов" % (s, len(subnet_to_hosts[s])))

    # ── 4. Запись файла ───────────────────────────────────────────────────────
    log.info("ШАГ 4/4 — Запись %s" % OUTPUT_FILE)

    wb = Workbook()

    # Лист 1: subnet_hosts
    ws1 = wb.active
    ws1.title = "subnet_hosts"
    hdr(ws1.cell(1, 1), "subnet",        C_HDR_GREEN)
    hdr(ws1.cell(1, 2), "hosts",         C_HDR_GREEN)
    hdr(ws1.cell(1, 3), "кол-во хостов", C_HDR_GREEN)
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 60
    ws1.column_dimensions["C"].width = 16
    ws1.row_dimensions[1].height = 22

    for row, s in enumerate(subnets, start=2):
        h_list = subnet_to_hosts[s]
        bg = C_HAS_HOSTS if h_list else C_NO_HOSTS
        dat(ws1.cell(row, 1), s,                                    bg=bg)
        dat(ws1.cell(row, 2), ", ".join(h_list) if h_list else "",  bg=bg)
        dat(ws1.cell(row, 3), len(h_list),                          bg=bg, bold=bool(h_list))

    log.info("Лист 'subnet_hosts' записан: %d строк" % len(subnets))

    # Лист 2: unique_hosts
    ws2 = wb.create_sheet("unique_hosts")
    hdr(ws2.cell(1, 1), "host (не входит ни в одну подсеть)", C_HDR_BLUE)
    ws2.column_dimensions["A"].width = 30
    ws2.row_dimensions[1].height = 22
    for i, h in enumerate(unique_hosts, start=2):
        dat(ws2.cell(i, 1), h, bg=C_ROW_EVEN if i % 2 == 0 else C_ROW_ODD)

    log.info("Лист 'unique_hosts' записан: %d строк" % len(unique_hosts))

    # Лист 3: unique_subnets
    ws3 = wb.create_sheet("unique_subnets")
    hdr(ws3.cell(1, 1), "subnet (нет ни одного хоста)", C_HDR_BLUE)
    ws3.column_dimensions["A"].width = 25
    ws3.row_dimensions[1].height = 22
    for i, s in enumerate(unique_subnets, start=2):
        dat(ws3.cell(i, 1), s, bg=C_ROW_EVEN if i % 2 == 0 else C_ROW_ODD)

    log.info("Лист 'unique_subnets' записан: %d строк" % len(unique_subnets))

    wb.save(OUTPUT_FILE)
    log.info("Файл сохранён: %s" % OUTPUT_FILE)
    log.info("=" * 60)
    log.info("Завершено успешно")
    log.info("=" * 60)

    print("\n✓ Готово! %s" % OUTPUT_FILE)
    print("  Лист 'subnet_hosts'   — %d подсетей (%d с хостами)" % (len(subnets), len(matched_subnets)))
    print("  Лист 'unique_hosts'   — %d хостов без подсети" % len(unique_hosts))
    print("  Лист 'unique_subnets' — %d подсетей без хостов" % len(unique_subnets))
    print("  Подробный лог:          %s" % LOG_FILE)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logging.getLogger("make_result").exception("КРИТИЧЕСКАЯ ОШИБКА: %s" % e)
        sys.exit(1)
