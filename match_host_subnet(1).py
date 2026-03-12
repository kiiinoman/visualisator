"""
Скрипт сопоставления хостов к основному файлу по подсети.

Логика:
  1. Читает result.xlsx  (колонки: host, subnet)
  2. Читает main.xlsx    (основной файл, колонка 'Подсеть')
  3. Вносит хосты в колонку L основного файла построчно:
       если подсеть строки main.xlsx совпадает с subnet из result.xlsx —
       хост(ы) записываются через запятую в колонку L этой строки.
  4. Хосты, чья подсеть не найдена в main.xlsx → лист 'Уникальные хосты'
  5. Подсети из main.xlsx, для которых нет ни одного хоста → лист 'Уникальные подсети'

Требования:
    pip install pandas openpyxl

Входные файлы (положить рядом со скриптом):
    - main.xlsx    — основной файл (колонка A = 'Подсеть', целевая колонка L)
    - result.xlsx  — результат предыдущего скрипта (колонки 'host', 'subnet')

Результат:
    - main_with_hosts.xlsx
"""

from typing import Optional, Dict, List, Set
import ipaddress
import collections

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string


# ── Настройки ─────────────────────────────────────────────────────────────────
MAIN_FILE    = "main.xlsx"             # основной файл (с колонкой Подсеть)
RESULT_FILE  = "result.xlsx"           # result от предыдущего скрипта
OUTPUT_FILE  = "main_with_hosts.xlsx"  # итоговый файл

SUBNET_COL_MAIN   = "Подсеть"         # название колонки подсети в main.xlsx
HOST_COL_RESULT   = "host"            # название колонки хоста  в result.xlsx
SUBNET_COL_RESULT = "subnet"          # название колонки подсети в result.xlsx
HOST_TARGET_COL   = "L"               # колонка для вставки хостов в main.xlsx
HOST_COL_HEADER   = "Хосты"           # заголовок колонки L
# ─────────────────────────────────────────────────────────────────────────────


# ── Стили ─────────────────────────────────────────────────────────────────────
COLOR_HEADER_MAIN = "2E7D32"
COLOR_HEADER_UNIQ = "1565C0"
COLOR_FILLED      = "C8E6C9"
COLOR_EMPTY       = "FFCDD2"
COLOR_WHITE       = "FFFFFF"
COLOR_LIGHT       = "F5F5F5"


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(cell, text, bg=COLOR_HEADER_MAIN):
    cell.value = text
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _data_cell(cell, bg=COLOR_WHITE):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _thin_border()
# ─────────────────────────────────────────────────────────────────────────────


def normalise_subnet(raw):
    # type: (str) -> Optional[str]
    try:
        return str(ipaddress.ip_network(raw.strip(), strict=False))
    except ValueError:
        return None


def build_subnet_host_map(result_df):
    # type: (pd.DataFrame) -> Dict[str, List[str]]
    mapping = collections.defaultdict(list)
    for _, row in result_df.iterrows():
        subnet_raw = str(row.get(SUBNET_COL_RESULT, "")).strip()
        host_val   = str(row.get(HOST_COL_RESULT, "")).strip()
        if not subnet_raw or not host_val:
            continue
        norm = normalise_subnet(subnet_raw)
        if norm:
            mapping[norm].append(host_val)
    return dict(mapping)


def find_col_index(ws, header_name, header_row=1):
    # type: (object, str, int) -> Optional[int]
    for cell in ws[header_row]:
        if cell.value and str(cell.value).strip() == header_name:
            return cell.column
    return None


def inject_hosts(ws, subnet_host_map, subnet_col_idx, host_col_idx, header_row=1):
    # type: (object, Dict, int, int, int) -> tuple
    matched_subnets = set()
    all_subnets_in_main = []

    _header_cell(ws.cell(row=header_row, column=host_col_idx), HOST_COL_HEADER)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_subnet = ws.cell(row=row_idx, column=subnet_col_idx).value
        if raw_subnet is None:
            continue
        raw_subnet = str(raw_subnet).strip()
        if not raw_subnet:
            continue

        norm = normalise_subnet(raw_subnet)
        all_subnets_in_main.append(norm if norm else raw_subnet)

        target_cell = ws.cell(row=row_idx, column=host_col_idx)

        if norm and norm in subnet_host_map:
            target_cell.value = ", ".join(subnet_host_map[norm])
            _data_cell(target_cell, bg=COLOR_FILLED)
            matched_subnets.add(norm)
        else:
            _data_cell(target_cell, bg=COLOR_EMPTY)

    ws.column_dimensions[get_column_letter(host_col_idx)].width = 35
    return matched_subnets, all_subnets_in_main


def write_unique_hosts_sheet(wb, result_df, matched_subnets):
    ws = wb.create_sheet("Уникальные хосты")
    _header_cell(ws.cell(row=1, column=1), "Хост", bg=COLOR_HEADER_UNIQ)
    _header_cell(ws.cell(row=1, column=2), "Подсеть из result", bg=COLOR_HEADER_UNIQ)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.row_dimensions[1].height = 22

    written = set()
    row_i = 2
    for _, row in result_df.iterrows():
        subnet_raw = str(row.get(SUBNET_COL_RESULT, "")).strip()
        host_val   = str(row.get(HOST_COL_RESULT, "")).strip()
        norm = normalise_subnet(subnet_raw)
        if norm not in matched_subnets and host_val not in written:
            bg = COLOR_LIGHT if row_i % 2 == 0 else COLOR_WHITE
            _data_cell(ws.cell(row=row_i, column=1, value=host_val), bg)
            _data_cell(ws.cell(row=row_i, column=2, value=subnet_raw), bg)
            written.add(host_val)
            row_i += 1
    return row_i - 2


def write_unique_subnets_sheet(wb, all_subnets_in_main, matched_subnets):
    ws = wb.create_sheet("Уникальные подсети")
    _header_cell(ws.cell(row=1, column=1), "Подсеть (нет хостов)", bg=COLOR_HEADER_UNIQ)
    ws.column_dimensions["A"].width = 25
    ws.row_dimensions[1].height = 22

    seen = set()
    row_j = 2
    for subnet_val in all_subnets_in_main:
        if subnet_val in matched_subnets or subnet_val in seen:
            continue
        seen.add(subnet_val)
        bg = COLOR_LIGHT if row_j % 2 == 0 else COLOR_WHITE
        _data_cell(ws.cell(row=row_j, column=1, value=subnet_val), bg)
        row_j += 1
    return row_j - 2


def main():
    print("=== Перенос хостов в основной файл ===\n")

    # 1. Загрузка result.xlsx
    print("[1/5] Загрузка хостов из '%s'..." % RESULT_FILE)
    result_df = pd.read_excel(RESULT_FILE, dtype=str)
    result_df.columns = result_df.columns.str.strip()
    for col in (HOST_COL_RESULT, SUBNET_COL_RESULT):
        if col not in result_df.columns:
            raise ValueError("Колонка '%s' не найдена в %s. Доступные: %s"
                             % (col, RESULT_FILE, list(result_df.columns)))
    result_df = result_df[
        result_df[SUBNET_COL_RESULT].notna() &
        (result_df[SUBNET_COL_RESULT].str.strip() != "")
    ]
    subnet_host_map = build_subnet_host_map(result_df)
    total_hosts = sum(len(v) for v in subnet_host_map.values())
    print("      Подсетей с хостами: %d, хостов всего: %d"
          % (len(subnet_host_map), total_hosts))

    # 2. Загрузка main.xlsx
    print("[2/5] Загрузка основного файла '%s'..." % MAIN_FILE)
    wb = load_workbook(MAIN_FILE)
    ws = wb.active

    header_row = 1
    subnet_col_idx = find_col_index(ws, SUBNET_COL_MAIN, header_row=1)
    if subnet_col_idx is None:
        subnet_col_idx = find_col_index(ws, SUBNET_COL_MAIN, header_row=2)
        header_row = 2
    if subnet_col_idx is None:
        headers_found = [c.value for c in ws[1]]
        raise ValueError("Колонка '%s' не найдена в %s. Заголовки строки 1: %s"
                         % (SUBNET_COL_MAIN, MAIN_FILE, headers_found))
    print("      Колонка '%s' -> %s"
          % (SUBNET_COL_MAIN, get_column_letter(subnet_col_idx)))

    host_col_idx = column_index_from_string(HOST_TARGET_COL)

    # 3. Вставка хостов
    print("[3/5] Вставка хостов в колонку %s..." % HOST_TARGET_COL)
    matched_subnets, all_subnets_in_main = inject_hosts(
        ws, subnet_host_map, subnet_col_idx, host_col_idx, header_row
    )
    print("      Строк заполнено: %d" % len(matched_subnets))

    # 4. Лист уникальных хостов
    print("[4/5] Формирование листов уникальных значений...")
    cnt_uh = write_unique_hosts_sheet(wb, result_df, matched_subnets)
    cnt_us = write_unique_subnets_sheet(wb, all_subnets_in_main, matched_subnets)

    # 5. Сохранение
    print("[5/5] Сохранение в '%s'..." % OUTPUT_FILE)
    wb.save(OUTPUT_FILE)

    print("\n✓ Готово! Файл: %s" % OUTPUT_FILE)
    print("  Основной лист          — хосты в колонке %s" % HOST_TARGET_COL)
    print("  'Уникальные хосты'     — %d запис(ей)" % cnt_uh)
    print("  'Уникальные подсети'   — %d запис(ей)" % cnt_us)


if __name__ == "__main__":
    main()
